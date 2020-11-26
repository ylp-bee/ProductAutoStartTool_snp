from Logger_module import logger
from ConfigManager import configs
from xml.dom import minidom
import random
import os
import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl


class Auto_write_agv_info():
    def __init__(self):
        self.excel_name = configs.excelpath
        self.sheet_name = 'Agv_info'
        self.title = ['id','type','ip','port','shell_port','layout_id','fts_port','simulation',
                      'pos_x','pos_y','pos_angle','jess_port','mac_addr']

        self.dockspath=os.path.join(configs.layoutpath,'docks.xml')
        self.need_get_agv_num=int(configs.agv_nums)
        self.start_agv_id=int(configs.start_agv_id)
        self.agv_type_id=int(configs.agv_type_id)

    def get_agv_info_from_docks(self):
        try:
            #打开xml文档
            dom = minidom.parse(self.dockspath)
            #得到文档元素对象
            root = dom.documentElement
            docks=root.getElementsByTagName('dock')
            choice_docks=random.sample(docks, int(self.need_get_agv_num))
            agv_info=[]
            for sd in choice_docks:
                #dock点
                docks_id=sd.getElementsByTagName("id")[0].childNodes[0].data
                docks_info=sd.getElementsByTagName("pos")[0]
                #x值、y值、坐标
                x=docks_info.getAttribute("x")
                y=docks_info.getAttribute("y")
                angle=docks_info.getAttribute("angle")
                single_dock_info=(docks_id,x,y,angle)
                # print(docks_id,x,y,angle)
                logger.info(f'get docks info:{docks_id},{x},{y},{angle}')
                agv_info.append(single_dock_info)
            return agv_info
        except Exception as e:
            logger.error('get_agv_info_from_docks error!原因：{},filename：{},line:{}'.format(e,e.__traceback__.tb_frame.f_globals['__file__'], e.__traceback__.tb_lineno))

    def Add_sheet(self,excel_name,sheet_name):
        try:
            wb = openpyxl.load_workbook(excel_name)
            wb.create_sheet(sheet_name, index=0)
            wb.save(excel_name)
        except Exception as e:
            logger.error('Add_sheet error!原因：{},filename：{},line:{}'.format(e,e.__traceback__.tb_frame.f_globals['__file__'], e.__traceback__.tb_lineno))

    def write_title_cell(self,excel_name,sheet_name,title_list):
        try:
            xfile = openpyxl.load_workbook(excel_name)
            sheet = xfile.get_sheet_by_name(sheet_name)
            for i in range(len(title_list)):
                #第1行第1列
                # print(i,title_list[i])
                sheet.cell(1, i+1).value =title_list[i]
            xfile.save(self.excel_name)
        except Exception as e:
            logger.error('write_title_cell error!原因：{},filename：{},line:{}'.format(e,e.__traceback__.tb_frame.f_globals['__file__'], e.__traceback__.tb_lineno))


    def check_sheet_and_clear_content(self,excel_name,sheet_name,title_list):
        try:
            isexists=os.path.exists(excel_name)
            if isexists:
                wb = xlrd.open_workbook(excel_name)
                sheet_names = wb.sheet_names()
                # print(f'sheet_names：{sheet_names},{self.sheet_name in sheet_names}')
                if(sheet_name not in sheet_names):
                    logger.info(f'current file：{excel_name} unexists sheet:{sheet_name},need create sheet！')
                    self.Add_sheet(excel_name, sheet_name)
                    self.write_title_cell(excel_name,sheet_name,title_list)
                else:
                    sheet = wb.sheet_by_name(self.sheet_name)
                    current_title = sheet.row_values(0)
                    current_rows = sheet.nrows
                    checkinfo = [xx for xx in title_list if xx not in current_title]
                    if (checkinfo != [] or current_rows>1 ):
                        logger.info(f'current file：{excel_name} title error or have content,detele sheet and add new！')
                        self.remove_sheet(excel_name,sheet_name)
                        self.Add_sheet(excel_name, sheet_name)
                        self.write_title_cell(excel_name, sheet_name, title_list)
        except Exception as e:
            logger.error('check_sheet_and_clear_content error!原因：{},filename：{},line:{}'.format(e,e.__traceback__.tb_frame.f_globals['__file__'], e.__traceback__.tb_lineno))

    def remove_sheet(self,excel_name,sheet_name):
        try:
            print(f'remove sheet:{sheet_name}')
            wb = openpyxl.load_workbook(excel_name)
            ws = wb[sheet_name]
            wb.remove(ws)
            wb.save(excel_name)
            print("It is over")
        except Exception as e:
            logger.error('remove_sheet error!原因：{},filename：{},line:{}'.format(e,e.__traceback__.tb_frame.f_globals['__file__'], e.__traceback__.tb_lineno))

    def write_agv_info_excel(self):
        try:
            get_info=self.get_agv_info_from_docks()
            xfile = openpyxl.load_workbook(self.excel_name)
            sheet = xfile.get_sheet_by_name(self.sheet_name)
            start_id=self.start_agv_id
            shell_port=int(configs.shell_port)+start_id
            port=10000+start_id
            fts_port=int(configs.fts_port)
            jess_port=int(configs.jess_port)
            for n in range(self.need_get_agv_num):
                self.title = ['id', 'type', 'ip', 'port', 'shell_port', 'layout_id', 'fts_port', 'simulation',
                              'pos_x', 'pos_y', 'pos_angle', 'jess_port', 'mac_addr']
                current_value=[start_id,self.agv_type_id,'127.0.0.1',port,shell_port,1,fts_port,1,get_info[n-1][1],get_info[n-1][2],get_info[n-1][3],jess_port,None]
                print(f'第{n+2}行数据：{current_value}')
                for i in range(len(self.title)):
                    sheet.cell(n+2, i+1).value = current_value[i]
                start_id=start_id+1
                port=port+1
                shell_port=shell_port+1
            xfile.save(self.excel_name)
        except Exception as e:
            logger.error('write_agv_info_excel error!原因：{},filename：{},line:{}'.format(e,e.__traceback__.tb_frame.f_globals['__file__'], e.__traceback__.tb_lineno))

    def excel_main(self):
        self.check_sheet_and_clear_content(self.excel_name,self.sheet_name,self.title)
        self.write_agv_info_excel()


